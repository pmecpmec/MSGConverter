"""
Excel Reader - Low-level Excel file handling

Deze module bevat utilities voor het lezen van Excel bestanden
met openpyxl en pandas.

Auteur: Pedro (met Cursor AI assistentie)
Datum: 4 februari 2026
"""

import logging
from pathlib import Path
from typing import List, Dict, Any, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd

logger = logging.getLogger(__name__)


class ExcelReader:
    """
    Low-level Excel reader utilities.

    Deze class biedt helper functies voor het lezen van Excel bestanden,
    sheets, ranges, etc. Gebruikt openpyxl voor cell-level operaties
    en pandas voor bulk data operaties.

    Example:
        >>> reader = ExcelReader()
        >>> sheets = reader.get_sheet_names(Path("file.xlsx"))
        >>> data = reader.read_range(Path("file.xlsx"), "Sheet1", "A1:D10")
    """

    def __init__(self):
        """Initialiseer de Excel reader."""
        logger.debug("ExcelReader geïnitialiseerd")

    def _validate_path(self, excel_path: Path) -> None:
        """Valideer dat het bestand bestaat en een Excel-bestand is."""
        if not excel_path.exists():
            raise FileNotFoundError(f"Bestand niet gevonden: {excel_path}")
        if excel_path.suffix.lower() not in (".xlsx", ".xlsm", ".xls"):
            raise ValueError(f"Ongeldig bestandstype: {excel_path.suffix}")

    def get_sheet_names(self, excel_path: Path) -> List[str]:
        """
        Haal alle sheet namen op uit een Excel bestand.

        Args:
            excel_path: Pad naar Excel bestand

        Returns:
            Lijst van sheet namen
        """
        self._validate_path(excel_path)
        logger.debug(f"Sheet namen ophalen van {excel_path}")

        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        try:
            return wb.sheetnames
        finally:
            wb.close()

    def read_sheet(
        self,
        excel_path: Path,
        sheet_name: Optional[str] = None,
        header_row: int = 1,
        max_rows: Optional[int] = None,
    ) -> List[Dict[str, Any]]:
        """
        Lees een volledige sheet als lijst van dictionaries.

        Elke rij wordt een dictionary met de header-namen als keys.

        Args:
            excel_path: Pad naar Excel bestand
            sheet_name: Naam van de sheet (None = eerste sheet)
            header_row: Rij nummer van de headers (1-based)
            max_rows: Maximum aantal data rijen om te lezen

        Returns:
            Lijst van dictionaries, één per data rij
        """
        self._validate_path(excel_path)
        logger.debug(f"Sheet '{sheet_name}' lezen van {excel_path}")

        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        try:
            ws = wb[sheet_name] if sheet_name else wb.active
            if ws is None:
                raise ValueError("Geen actieve sheet gevonden")

            rows = list(ws.iter_rows(values_only=True))
            if not rows or len(rows) < header_row:
                logger.warning("Sheet is leeg of heeft te weinig rijen")
                return []

            # Headers ophalen (genormaliseerd: strip, upper)
            raw_headers = rows[header_row - 1]
            headers = []
            for h in raw_headers:
                if h is not None:
                    headers.append(str(h).strip().upper())
                else:
                    headers.append(None)

            # Data rijen verwerken
            data_rows = rows[header_row:]
            if max_rows:
                data_rows = data_rows[:max_rows]

            result = []
            for row_idx, row in enumerate(data_rows, start=header_row + 1):
                # Skip volledig lege rijen
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue

                record = {}
                for col_idx, header in enumerate(headers):
                    if header is None:
                        continue
                    value = row[col_idx] if col_idx < len(row) else None
                    record[header] = value

                record["_row_number"] = row_idx
                result.append(record)

            logger.debug(f"{len(result)} data rijen gelezen")
            return result
        finally:
            wb.close()

    def read_range(
        self,
        excel_path: Path,
        sheet_name: str,
        cell_range: str
    ) -> List[List[Any]]:
        """
        Lees een cell range uit een Excel sheet.

        Args:
            excel_path: Pad naar Excel bestand
            sheet_name: Naam van de sheet
            cell_range: Excel range notation (bijv. "A1:D10")

        Returns:
            2D lijst met cell waarden
        """
        self._validate_path(excel_path)
        logger.debug(f"Range {cell_range} lezen van {sheet_name}")

        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        try:
            ws = wb[sheet_name]
            result = []
            for row in ws[cell_range]:
                result.append([cell.value for cell in row])
            return result
        finally:
            wb.close()

    def find_header_row(
        self,
        excel_path: Path,
        sheet_name: str,
        expected_headers: List[str]
    ) -> Optional[int]:
        """
        Vind de rij met headers in een Excel sheet.

        Zoekt in de eerste 20 rijen naar een rij die minstens de helft
        van de verwachte headers bevat.

        Args:
            excel_path: Pad naar Excel bestand
            sheet_name: Naam van de sheet
            expected_headers: Lijst van verwachte header namen

        Returns:
            Rij nummer (1-based) of None als niet gevonden
        """
        self._validate_path(excel_path)
        logger.debug(f"Headers zoeken in {sheet_name}")

        expected_upper = {h.upper() for h in expected_headers}
        threshold = len(expected_upper) / 2

        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        try:
            ws = wb[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
                row_values = {
                    str(cell).strip().upper()
                    for cell in row
                    if cell is not None
                }
                matches = row_values & expected_upper
                if len(matches) >= threshold:
                    logger.info(f"Headers gevonden op rij {row_idx}: {matches}")
                    return row_idx

            logger.warning("Geen header rij gevonden")
            return None
        finally:
            wb.close()

    def read_as_dataframe(
        self,
        excel_path: Path,
        sheet_name: Optional[str] = None,
        header_row: int = 0,
    ) -> "pd.DataFrame":
        """
        Lees een sheet als pandas DataFrame.

        Args:
            excel_path: Pad naar Excel bestand
            sheet_name: Naam van de sheet (None = eerste sheet)
            header_row: Rij index van de headers (0-based voor pandas)

        Returns:
            pandas DataFrame
        """
        self._validate_path(excel_path)
        logger.debug(f"DataFrame lezen van {excel_path}")

        df = pd.read_excel(
            excel_path,
            sheet_name=sheet_name or 0,
            header=header_row,
            engine="openpyxl",
        )

        # Normaliseer kolomnamen
        df.columns = [str(c).strip().upper() for c in df.columns]

        logger.debug(f"DataFrame: {len(df)} rijen, {len(df.columns)} kolommen")
        return df
